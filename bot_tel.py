from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from config import TOKEN
import pandas as pd
import os
import urlextract
import openpyxl
from openpyxl import Workbook
from datetime import date
import datetime
import calendar
import schedule
import time
from openpyxl import load_workbook

bot = Bot(TOKEN)
dp = Dispatcher(bot)



extractor = urlextract.URLExtract()
wb = Workbook()
ws = wb.active


@dp.message_handler(commands=['start'])
async def command_start(message : types.Message):
	await message.answer("Привет. Я чат бот WEBLAB420, мне ты можешь отправлять ссылки на свои работы" + \
		' (Google driver, Yandex disk, Tilda, Figma, Behance, Dribbble).\n\nНе забывай о дедлайнах! Проекты,' + \
		' которые ты пришлешь после указанной даты - будут аннулированы.\n\n Ссылка на чат:... ')



@dp.message_handler()
async def echo_send(message : types.Message):
	valid = message.text
	urls = extractor.find_urls(valid)

	
	if not urls:
		await message.answer('Извини, я принимаю только ссылки')
	else:
		wz = openpyxl.Workbook()

		now = datetime.datetime.now()
		my_date = date.today()
		day_week = calendar.day_name[my_date.weekday()]
		date_now = 'sheet '+str(now.month)+'-'+str(now.day)

		if day_week == 'Wednesday':
			wz = load_workbook(filename = 'work.xlsx')
			if date_now in wz.sheetnames:
				pass
			else: 
				wz.create_sheet(title = date_now, index = 0)
				wz.save("work.xlsx")
		

		text = ''.join(urls)
		await message.answer('Спасибо, я принял твой проект')
		user = message.from_user.username
		sheet = wz[date_now]
		df = pd.DataFrame([[text, user, now]],
				   index=['row 0'],
				   columns=['col 1', 'col 2', 'col 3'])
		with pd.ExcelWriter("work.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
			df.to_excel(writer, sheet_name=date_now,header=None, startrow=writer.sheets[date_now].max_row,index=False)


	
		

def messege_sunday():
	bot.send_message(message.chat.id, 'Привет! Я сообщение, отправленное в 7 часов утра.')


schedule.every().sunday.at("23:00").do(messege_sunday)
while True:
	schedule.run_pending()
	time.sleep(1)
	executor.start_polling(dp, skip_updates=True)

